# load workbook and load_workbook
from openpyxl import Workbook, load_workbook
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.compat import range
from openpyxl.worksheet import Worksheet, ColumnDimension, RowDimension
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import re
from datetime import date, timedelta
import datetime

# define natural sorting algorithm
def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split('(\d+)', text) ]
 
# find column number based on provided workbook and title
def findcolumn(workbook, desired_title):
	data = workbook.get_sheet_by_name('data')
	end_column = data.get_highest_column() + 1
	for j in range(1, end_column):
		excel_title = str(data.cell(row = 1, column = j).value)
		if excel_title == desired_title:
			return j

# make 'data' the active sheet for the PO data manipulation
def activesheet(workbook):
	data = workbook.get_sheet_by_name('data')

# function to find the maximum row height
def maxrow(workbook):
	maxrow_book = workbook.get_sheet_by_name('data')
	end_row = int(maxrow_book.get_highest_row())
	return end_row
	
# Create list of all projects then make new workbook with sheets
# specific to each project, in order
def listproj():
	global wb
	wb = Workbook()
	PO_Project_Col = findcolumn(po, 'Project')
	po_data = po.get_sheet_by_name('data')
	end_row = int(po_data.get_highest_row())
	req_data = req.get_sheet_by_name('data')
	end_row_req = int(req_data.get_highest_row())
	global Projects
	Projects = []
	for i in range(2, end_row):
		if str(po_data.cell(column = PO_Info[10], row = i).value) != 'Yes':
			Projects.append(str(po_data.cell(row = i, column = PO_Project_Col).value))
	for j in range(2, end_row_req):
		Projects.append(str(req_data.cell(row = j, column = REQ_Info[0]).value))
	Projects = list(set(filter(None, Projects)))
	Projects.sort(key=natural_keys)
	Size_Project_List = len(Projects)
	p = {}
	for i in range(0, Size_Project_List):
		z= i + 1
		p['proj_sheet_%02d' % z ] = wb.create_sheet(title = Projects[i])
	sheetName = 'Sheet'
	ws = wb.get_sheet_by_name(sheetName)
	wb.remove_sheet(ws)

# create column titles in new sheet, 9 total. 
# adds in status column at the end
# and also creates an even width among the columns
def new_column_titles():
	for i in range(0, len(Projects)): 
		ws = wb.get_sheet_by_name(Projects[i])
		for x in range(0, len(PO_Titles)):
			ws.cell(column = (x + 1), row = 1, value = "%s" % PO_Titles[x]).font = Font(bold = True)

# PO transfer function. 
def transfer_po():
	po_data = po.get_sheet_by_name('data')
	end_row_po = int(po_data.get_highest_row()) + 1
	Counter = [1]*len(Projects)
	for x in range(0, len(Projects)):
		for j in range(2, end_row_po):
			if str(po_data.cell(column = PO_Project_Col, row = j).value) == Projects[x] and str(po_data.cell(column = PO_Info[10], row = j).value) != 'Yes':
				ws = wb.get_sheet_by_name(Projects[x])
				ws.cell(column = (PO_Titles.index('PO') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[0], row = j).value
				ws.cell(column = (PO_Titles.index('REQ') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[1], row = j).value
				ws.cell(column = (PO_Titles.index('Pos.') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[2], row = j).value
				ws.cell(column = (PO_Titles.index('Activity') + 1), row = Counter[x] + 1).value = str(po_data.cell(column = PO_Info[3], row = j).value)
				ws.cell(column = (PO_Titles.index('Part Number') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[4], row = j).value
				ws.cell(column = (PO_Titles.index('Part Description') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[5], row = j).value
				ws.cell(column = (PO_Titles.index('Business Partner') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[6], row = j).value
				ws.cell(column = (PO_Titles.index('Order') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[7], row = j).value
				ws.cell(column = (PO_Titles.index('Unit') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[8], row = j).value
				ws.cell(column = (PO_Titles.index('RECV') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[9], row = j).value
				ws.cell(column = (PO_Titles.index('Receipt Date') + 1), row = Counter[x] + 1).value = po_data.cell(column = PO_Info[11], row = j).value
				ws.cell(column = (PO_Titles.index('Receipt Date') + 1), row = Counter[x] + 1).number_format = 'MM/DD/YYYY'
				Counter[x] += 1					

# transfer all newly created reqs to POREQ workbook
def transfer_req():
	req_data = req.get_sheet_by_name('data')
	end_row_req = int(req_data.get_highest_row()) + 1
	for k in range (2, end_row_req):
		if str(req_data.cell(column = REQ_Info[1], row = k).value) == 'Created':
			New_Req_Project = str(req_data.cell(column = REQ_Info[0], row = k).value)
			ws = wb.get_sheet_by_name(New_Req_Project)
			ws_top = int(ws.get_highest_row())
			ws.cell(column = (PO_Titles.index('REQ') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[2], row = k).value
			ws.cell(column = (PO_Titles.index('Pos.') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[9], row = k).value
			ws.cell(column = (PO_Titles.index('Activity') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[3], row = k).value
			ws.cell(column = (PO_Titles.index('Part Number') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[4], row = k).value
			ws.cell(column = (PO_Titles.index('Part Description') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[5], row = k).value
			ws.cell(column = (PO_Titles.index('Business Partner') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[6], row = k).value
			ws.cell(column = (PO_Titles.index('Order') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[7], row = k).value
			ws.cell(column = (PO_Titles.index('Unit') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[8], row = k).value	
			ws.cell(column = (PO_Titles.index('Receipt Date') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[10], row = k).value
			ws.cell(column = (PO_Titles.index('Receipt Date') + 1), row = (ws_top + 1)).number_format = 'MM/DD/YYYY'
			ws.cell(column = (PO_Titles.index('Status') + 1), row = (ws_top) + 1).value = 'Req Created'
		if str(req_data.cell(column = REQ_Info[1], row = k).value) == 'Pending Approval':
			New_Req_Project = str(req_data.cell(column = REQ_Info[0], row = k).value)
			ws = wb.get_sheet_by_name(New_Req_Project)
			ws_top = int(ws.get_highest_row())
			ws.cell(column = (PO_Titles.index('REQ') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[2], row = k).value
			ws.cell(column = (PO_Titles.index('Pos.') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[9], row = k).value
			ws.cell(column = (PO_Titles.index('Activity') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[3], row = k).value
			ws.cell(column = (PO_Titles.index('Part Number') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[4], row = k).value
			ws.cell(column = (PO_Titles.index('Part Description') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[5], row = k).value
			ws.cell(column = (PO_Titles.index('Business Partner') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[6], row = k).value
			ws.cell(column = (PO_Titles.index('Order') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[7], row = k).value
			ws.cell(column = (PO_Titles.index('Unit') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[8], row = k).value	
			ws.cell(column = (PO_Titles.index('Receipt Date') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[10], row = k).value
			ws.cell(column = (PO_Titles.index('Receipt Date') + 1), row = (ws_top + 1)).number_format = 'MM/DD/YYYY'
			ws.cell(column = (PO_Titles.index('Status') + 1), row = (ws_top) + 1).value = 'Req Pending Approval'
		if str(req_data.cell(column = REQ_Info[1], row = k).value) == 'Approved':
			New_Req_Project = str(req_data.cell(column = REQ_Info[0], row = k).value)
			ws = wb.get_sheet_by_name(New_Req_Project)
			ws_top = int(ws.get_highest_row())
			ws.cell(column = (PO_Titles.index('REQ') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[2], row = k).value
			ws.cell(column = (PO_Titles.index('Pos.') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[9], row = k).value
			ws.cell(column = (PO_Titles.index('Activity') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[3], row = k).value
			ws.cell(column = (PO_Titles.index('Part Number') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[4], row = k).value
			ws.cell(column = (PO_Titles.index('Part Description') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[5], row = k).value
			ws.cell(column = (PO_Titles.index('Business Partner') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[6], row = k).value
			ws.cell(column = (PO_Titles.index('Order') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[7], row = k).value
			ws.cell(column = (PO_Titles.index('Unit') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[8], row = k).value	
			ws.cell(column = (PO_Titles.index('Receipt Date') + 1), row = (ws_top + 1)).value = req_data.cell(column = REQ_Info[10], row = k).value
			ws.cell(column = (PO_Titles.index('Receipt Date') + 1), row = (ws_top + 1)).number_format = 'MM/DD/YYYY'
			ws.cell(column = (PO_Titles.index('Status') + 1), row = (ws_top) + 1).value = 'Req Approved'

# depeneding on the PO or REQ, a status name and color indicator are updated
def color_status():
	greenFill = PatternFill(fill_type = 'solid', start_color = '66FF66', end_color = '66FF66')
	yellowFill = PatternFill(fill_type = 'solid', start_color = 'FFFF66', end_color = 'FFFF66')
	redFill = PatternFill(fill_type = 'solid', start_color = 'FF6666', end_color = 'FF6666')
	orangeFill = PatternFill(fill_type = 'solid', start_color = 'FAAC58', end_color = 'FAAC58')
	blueFill = PatternFill(fill_type = 'solid', start_color = '81F7F3', end_color = '81F7F3')
	for x in range(0, len(Projects)):
		ws = wb.get_sheet_by_name(Projects[x])
		end_row_ws = ws.get_highest_row() + 1
		for j in range(2, end_row_ws):
			if ws.cell(column = (PO_Titles.index('Order') + 1), row = j).value == ws.cell(column = (PO_Titles.index('RECV') + 1), row = j).value:
				ws.cell(column = (PO_Titles.index('Status') + 1), row = j).value = 'Received'
				ws.cell(column = (PO_Titles.index('Status') + 1), row = j).fill = greenFill
			if ws.cell(column = (PO_Titles.index('PO') + 1), row = j).value != None and ws.cell(column = (PO_Titles.index('Order') + 1), row = j).value != ws.cell(column = (PO_Titles.index('RECV') + 1), row = j).value:
				ws.cell(column = (PO_Titles.index('Status') + 1), row = j).value = 'on PO'
				ws.cell(column = (PO_Titles.index('Status') + 1), row = j).fill = yellowFill
			if str(ws.cell(column = (PO_Titles.index('Status') + 1), row = j).value) == 'Req Approved':
				ws.cell(column = (PO_Titles.index('Status') + 1), row = j).fill = redFill
			if str(ws.cell(column = (PO_Titles.index('Status') + 1), row = j).value) == 'Req Pending Approval':
				ws.cell(column = (PO_Titles.index('Status') + 1), row = j).fill = orangeFill
			if str(ws.cell(column = (PO_Titles.index('Status') + 1), row = j).value) == 'Req Created':
				ws.cell(column = (PO_Titles.index('Status') + 1), row = j).fill = blueFill

# compare planned receipt date of req vs. po
def time_comp():
	req_data = req.get_sheet_by_name('data')
	end_row_req = int(req_data.get_highest_row())
	for x in range (2, end_row_req + 1):
		for i in range(0, len(Projects)):
			ws = wb.get_sheet_by_name(Projects[i])
			end_row_test = ws.get_highest_row()
			for j in range(2, end_row_test + 1):
				if str(req_data.cell(column = REQ_Info[2], row = x).value) == str(ws.cell(column = (PO_Titles.index('REQ') + 1), row = j).value) and str(req_data.cell(column = REQ_Info[5], row = x).value) == str(ws.cell(column = (PO_Titles.index('Part Description') + 1), row = j).value):
						a = str(ws.cell(column = (PO_Titles.index('Receipt Date') + 1), row = j).value)
						new_a = datetime.datetime.strptime(a, '%Y-%m-%d %H:%M:%S')
						b = str(req_data.cell(column = REQ_Info[10], row = x).value)
						new_b = datetime.datetime.strptime(b, '%Y-%m-%d %H:%M:%S')
						diff = new_a - new_b
						ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).value = diff.days

def time_comp_correction():
	greenFill = PatternFill(fill_type = 'solid', start_color = '66FF66', end_color = '66FF66')
	redFill = PatternFill(fill_type = 'solid', start_color = 'FAAC58', end_color = 'FAAC58')
	noneFill = PatternFill(fill_type = None , start_color = 'FFFFFF', end_color = 'FFFFFF')
	for x in range(0, len(Projects)):
		ws = wb.get_sheet_by_name(Projects[x])
		end_row_test = ws.get_highest_row()
		for j in range (2, end_row_test + 1):
			if ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).value == -1 or ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).value == 0 or ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).value == None:
				ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).value = ""
				ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).fill = noneFill
			elif ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).value >= 7 and ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).value != "":
				ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).fill = redFill
			elif ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).value <= -7:
				ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).fill = greenFill
			else: 
				ws.cell(column = (PO_Titles.index('REQ Date') + 1), row = j).fill = noneFill
	
				
def auto_width():
	for x in range(0, len(Projects)):
		ws = wb.get_sheet_by_name(Projects[x])
		end_column_ws = ws.get_highest_column()
		end_row_ws = ws.get_highest_row()
		column_widths = [0] * (end_column_ws)
		for i in range (1, end_column_ws + 1):
			for j in range(1, end_row_ws + 1):
				if len(str((ws.cell(column = i, row = j)).value)) > column_widths[i - 1]:
					column_widths[i-1] = len(str(ws.cell(column = i, row = j).value))
			ws.column_dimensions[get_column_letter(i)].width = (column_widths[i - 1] + 3)
		ws.column_dimensions[get_column_letter(PO_Titles.index('Receipt Date') + 1)].width = 12

# add filter to each column
def auto_filter():
	for x in range (0, len(Projects)):
		ws = wb.get_sheet_by_name(Projects[x])
		end_column_ws = ws.get_highest_column()
		end_row_ws = ws.get_highest_column()
		ws.auto_filter.ref = 'A1:%s%s' % (get_column_letter(end_column_ws), end_row_ws)
	
# fill two columns, the total inevntory on hand and locations
def inventory():
	inv_data = inv.get_sheet_by_name('data')
	end_row_inv = int(inv_data.get_highest_row())
	for x in range(0, len(Projects)):
		ws = wb.get_sheet_by_name(Projects[x])
		end_row_test = int(ws.get_highest_row())
		for j in range(2, end_row_test + 1):
			item_list = []
			for k in range(2, end_row_inv + 1):
				if ws.cell(column = (PO_Titles.index('Part Number') + 1), row = j).value == inv_data.cell(column = INV_Info[0], row = k).value and inv_data.cell(column = INV_Info[4], row = k).value > 0:
					item_list.extend([k])
			if len(item_list) > 0:
				item_on_hand = []
				item_locations = []
				item_projects = []
				Locations = "%s:%s = %d, "
				for p in range (0, len(item_list)):
					item_on_hand.extend([int(inv_data.cell(column = INV_Info[4], row = item_list[p]).value)])
					item_locations.extend([str(inv_data.cell(column = INV_Info[3], row = item_list[p]).value)])
					item_projects.extend([str(inv_data.cell(column = INV_Info[2], row = item_list[p]).value)])
				ws.cell(column = (PO_Titles.index('On Hand') + 1), row = j).value = sum(item_on_hand)
				Locations_total = Locations*len(item_list)
				Locations_fixed = Locations_total[:-2]
				locations_and_qty = [None]*(len(item_on_hand) + len(item_locations) + len(item_projects))
				locations_and_qty[::3] = item_projects
				locations_and_qty[1::3] = item_locations
				locations_and_qty[2::3] = item_on_hand
				ws.cell(column = (PO_Titles.index('Locations') + 1), row = j).value = Locations_fixed % tuple(locations_and_qty)
			else:
				ws.cell(column = (PO_Titles.index('On Hand') + 1), row = j).value = int(0)

# use if needed project specific extra parts, currently configured to scan all projects for 'on-hand' quntities		
# and inv_data.cell(column = INV_Info[2], row = k).value == Projects[x] 			

			
		

### start program ###
# load po and req workbook
# new_filename = raw_input("Todays Date? (MM/DD/YY) ")
new_filename = raw_input("New File Name: ")
# new_filename = 'test.xlsx'
po_location = raw_input(r"PO Info Location: ")
req_location = raw_input(r"REQ Info Location: ")
inv_location = raw_input(r"Inventory Info Location: ")

po = load_workbook(filename = po_location.lower())
req = load_workbook(filename = req_location.lower())
inv = load_workbook(filename = inv_location.lower())

# po = load_workbook('c:\\Users\crussell\\Documents\\polines.xlsx')
# req = load_workbook('c:\\Users\\crussell\\Documents\\reqlines.xlsx')
# inv = load_workbook('c:\\Users\crussell\\Documents\\inventorylines.xlsx')

# defining all significant title locations among both the PO
# and REQ workbooks. Create one list for all important columns
# in the PO workbook
PO_Position_Col = findcolumn(po, 'Order Line') + 1
PO_Project_Col = findcolumn(po, 'Project')
PO_Purchase_Order_Col = findcolumn(po, 'Order Line')
PO_Purchase_Req_Col = findcolumn(po, 'Requisition')
PO_Activity_Col = findcolumn(po, 'Activity')
PO_Part_Number_Col = findcolumn(po, 'Item') + 1
PO_Part_Description_Col = findcolumn(po, 'Description')
PO_Business_Partner_Name_Col = findcolumn(po, 'Ship-from BP') + 1
PO_Ordered_Quantity_Col = findcolumn(po, 'Ordered Quantity')
PO_Ordered_Units_Col = PO_Ordered_Quantity_Col + 1
PO_Received_Quantity_Col = findcolumn(po, 'Received Quantity')
PO_Deleted = findcolumn(po, 'Canceled')
PO_Planned_Receipt_Date = findcolumn(po, 'Planned Receipt Date')


# list of all title column positions in the PO file
PO_Info = [PO_Purchase_Order_Col, PO_Purchase_Req_Col, PO_Position_Col, PO_Activity_Col,
PO_Part_Number_Col, PO_Part_Description_Col, PO_Business_Partner_Name_Col,
PO_Ordered_Quantity_Col, PO_Ordered_Units_Col, PO_Received_Quantity_Col, PO_Deleted, PO_Planned_Receipt_Date]

# list of all title names
PO_Titles =['PO', 'REQ', 'Pos.', 'Activity', 'Part Number', 'Part Description',
'Receipt Date', 'REQ Date', 'Business Partner', 'Order', 'Unit', 'RECV', 'Status', 'On Hand', 'Locations']


# defining important REQ variables and lists
REQ_Status = findcolumn(req, 'Status')
REQ_Project_Col = findcolumn(req, 'Project')
REQ_Purchase_Req_Col = findcolumn(req, 'Requisition')
REQ_Activity_Col = findcolumn(req, 'Activity')
REQ_Part_Number_Col = findcolumn(req, 'Item') + 1
REQ_Part_Description_Col = findcolumn(req, 'Item Description')
REQ_Business_Partner_Name_Col = findcolumn(req, 'Business Parter Description')
REQ_Ordered_Quantity_Col = findcolumn(req, 'Order Quantity')
REQ_Ordered_Units_Col = REQ_Ordered_Quantity_Col + 1
REQ_Position = findcolumn(req, 'Position')
REQ_Planned_Receipt_Date = findcolumn(req, 'Requested Date')


# list of all column name positions in the REQ file
REQ_Info = [REQ_Project_Col, REQ_Status, REQ_Purchase_Req_Col, REQ_Activity_Col,
REQ_Part_Number_Col, REQ_Part_Description_Col, REQ_Business_Partner_Name_Col,
REQ_Ordered_Quantity_Col, REQ_Ordered_Units_Col, REQ_Position, REQ_Planned_Receipt_Date]
listproj()

# defining important INV variables and lists
INV_Part_Number_Col = findcolumn(inv, 'Item') + 1
INV_Part_Description_Col = findcolumn(inv, 'Item') + 2
INV_Project_Col = findcolumn(inv, 'Project')
INV_Activity_Col = findcolumn(inv, 'Activity')
INV_Inventory_On_Hand_Col = findcolumn(inv, 'Inventory on Hand')

#list of all column name positions in the inv file
INV_Info = [INV_Part_Number_Col, INV_Part_Description_Col, INV_Project_Col, INV_Activity_Col, INV_Inventory_On_Hand_Col]

new_column_titles()

transfer_po()

transfer_req()

inventory()

color_status()

time_comp()

time_comp_correction()

auto_width()

auto_filter()

wb.save(filename = new_filename)

k = raw_input("Press enter to close the program")

			

	
	

